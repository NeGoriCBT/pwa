from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from typing import List, Dict
import os

def generate_excel(entries: List[Dict], user_id: int, entry_type: str = 'thoughts') -> str:
    wb = Workbook()
    ws = wb.active
    
    if entry_type == 'exposure':
        ws.title = "Дневник экспозиций"
        headers = [
            'Название ситуации', 'Дата и время события',
            'Ожидания', 'Реальность', 'Статус'
        ]
    else:
        ws.title = "Дневник мыслей"
        headers = [
            'Дата и время', 'Ситуация', 'Эмоции (до)', 
            'Автом. мысль (уверенность%)', 'Действие', 'Доводы За', 
            'Доводы Против', 'Альтерн. мысли (уверенность%)', 
            'Эмоции (после)', 'Комментарий'
        ]
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Записываем данные
    for row_idx, entry in enumerate(entries, 2):
        if entry_type == 'exposure':
            status = "✅ Завершено" if entry.get('reality_received', 0) else "⏳ Ожидает"
            data = [
                entry.get('situation_name', ''),
                entry.get('event_datetime', ''),
                entry.get('expectations', ''),
                entry.get('reality', 'Еще не заполнено'),
                status
            ]
        else:
            # Эмоции до
            emotions_before = []
            for em in entry.get('emotions_before', []):
                emotions_before.append(f"{em.get('emotion')}: {em.get('intensity')}%")
            emotions_before_str = '; '.join(emotions_before)
            
            # Автоматическая мысль
            auto_thought = entry.get('automatic_thought', '')
            auto_confidence = entry.get('automatic_thought_confidence', 0)
            auto_thought_str = f"{auto_thought} ({auto_confidence}%)"
            
            # Альтернативные мысли
            alt_thoughts = []
            for alt in entry.get('alternative_thoughts', []):
                alt_thoughts.append(f"{alt.get('thought')} ({alt.get('confidence')}%)")
            alt_thoughts_str = '; '.join(alt_thoughts)
            
            # Эмоции после
            emotions_after = []
            for em in entry.get('emotions_after', []):
                emotions_after.append(f"{em.get('emotion')}: {em.get('intensity')}%")
            emotions_after_str = '; '.join(emotions_after)
            
            data = [
                entry.get('timestamp'),
                entry.get('situation', ''),
                emotions_before_str,
                auto_thought_str,
                entry.get('action', ''),
                entry.get('evidence_for', ''),
                entry.get('evidence_against', ''),
                alt_thoughts_str,
                emotions_after_str,
                entry.get('note_to_future_self', '')
            ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Сохраняем файл
    if entry_type == 'exposure':
        filename = f"exposures_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        filename = f"cognitive_diary_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('temp', filename)
    os.makedirs('temp', exist_ok=True)
    wb.save(filepath)
    
    return filepath